import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

# 记录：(日期, 年份, 说明, 金额)
records = [
    # 2023年
    ("2023-12-09", "2023", "存入", 37000),
    ("2023-12-15", "2023", "存入", 17500),
    # 2024年
    ("2024-01-15", "2024", "存入", 5500),
    ("2024-02-20", "2024", "存入", 10340),
    ("2024-03-15", "2024", "存入", 12660),
    ("2024-04-15", "2024", "存入", 11000),
    ("2024-04-22", "2024", "支出：空调热水器", -2280),
    ("2024-04-24", "2024", "支出：妙头儿买衣物", -1000),
    ("2024-04-25", "2024", "支出：买燕窝两盒", -2217),
    ("2024-04-30", "2024", "支出5000实际3000加回2000", -3000),
    ("2024-05-08", "2024", "存入：公积金提取", 50000),
    ("2024-05-15", "2024", "存入", 7700),
    ("2024-05-29", "2024", "存入：年终奖", 65000),
    ("2024-06-15", "2024", "存入（妙存3000）", 13200),
    ("2024-07-17", "2024", "存入", 9400),
    ("2024-07-28", "2024", "支出：妙头打针", -1331),
    ("2024-08-16", "2024", "存入", 9000),
    ("2024-08-xx", "2024", "支出：金饰（手镯+戒指+猪+耳环）", -75307),
    ("2024-09-15", "2024", "存入", 8387),
    ("2024-10-10", "2024", "支出：房租", -3660),
    ("2024-10-16", "2024", "存入（妙存10000）", 10000),
    ("2024-10-16", "2024", "支出：买妙衣服", -1835),
    ("2024-11-19", "2024", "存入", 10000),
    ("2024-12-24", "2024", "存入", 10000),
    # 2025年
    ("2025-01-19", "2025", "支出：房契税+妙买衣服", -14117),
    ("2025-02-17", "2025", "存入：过年结算", 6800),
    ("2025-03-16", "2025", "存入", 10000),
    ("2025-04-16", "2025", "存入", 10000),
    ("2025-05-18", "2025", "存入", 8000),
    ("2025-06-08", "2025", "支出：虎家+封窗尾款", -26500),
    ("2025-06-15", "2025", "存入", 1000),
    ("2025-06-xx", "2025", "支出：瓷砖+全屋定制定金", -6273),
    ("2025-07-18", "2025", "存入（含支出橱柜+全屋定制）", 83481),
    ("2025-08-04", "2025", "支出：美缝定金+装修电器", -16966),
    ("2025-09-13", "2025", "支出：婚备", -40000),
    ("2025-09-23", "2025", "支出：婚礼用", -8189),
    ("2025-09-23", "2025", "存入：婚礼礼金+妙+刚", 52143),
    ("2025-10-15", "2025", "存入", 8000),
    ("2025-11-15", "2025", "支出：装修", -10000),
    ("2025-12-16", "2025", "存入", 8000),
    # 2026年
    ("2026-01-08", "2026", "支出：布布还花呗", -500),
    ("2026-01-25", "2026", "支出：买银", -27030),
    ("2026-02-03", "2026", "支出：布布还信用卡花呗", -6000),
    ("2026-02-12", "2026", "支出：过年红包+还花呗", -1300),
    ("2026-02-21", "2026", "存入", 6500),
    ("2026-02-27", "2026", "存入（妙3000+刚330）", 3330),
    ("2026-03-17", "2026", "存入", 8000),
    # 补齐差额
    ("2026-03-24", "2026", "其他（补齐差额）", -66436),
]

# 初始余额：妙一二83000 + 刚布布206000 = 289000
balance = 111000
rows = []
for date, year, desc, change in records:
    balance += change
    rows.append((date, year, desc, change, balance))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "存钱流水"

# 样式
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
income_fill = PatternFill('solid', start_color='E2EFDA')
expense_fill = PatternFill('solid', start_color='FCE4D6')
other_fill = PatternFill('solid', start_color='FFF2CC')
center = Alignment(horizontal='center', vertical='center')
right = Alignment(horizontal='right', vertical='center')
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ['日期', '年份', '说明', '收支金额（元）', '累计余额（元）']
ws.append(headers)
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = PatternFill('solid', start_color='2E75B6')
    cell.alignment = center
    cell.border = border

for i, (date, year, desc, change, bal) in enumerate(rows, 2):
    if "其他" in desc:
        row_fill = other_fill
    elif change > 0:
        row_fill = income_fill
    else:
        row_fill = expense_fill

    ws.cell(row=i, column=1, value=date).alignment = center
    ws.cell(row=i, column=2, value=year).alignment = center
    ws.cell(row=i, column=3, value=desc)
    ws.cell(row=i, column=4, value=change)
    ws.cell(row=i, column=5, value=bal)

    for col in range(1, 6):
        cell = ws.cell(row=i, column=col)
        cell.fill = row_fill
        cell.border = border
        if col in (4, 5):
            cell.number_format = '#,##0'
            cell.alignment = right
        elif col in (1, 2):
            cell.alignment = center

ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 34
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 18
ws.freeze_panes = 'A2'

# 汇总
ws2 = wb.create_sheet("汇总统计")
total_income = sum(r[3] for r in rows if r[3] > 0)
total_expense = sum(r[3] for r in rows if r[3] < 0)
final = rows[-1][4]

summary = [
    ("起始余额", 111000),
    ("总存入", total_income),
    ("总支出", abs(total_expense)),
    ("当前余额", final),
    ("净增加", final - 111000),
]
ws2.append(["项目", "金额（元）"])
for cell in ws2[1]:
    cell.font = header_font
    cell.fill = PatternFill('solid', start_color='2E75B6')
    cell.alignment = center
    cell.border = border
for item, val in summary:
    ws2.append([item, val])
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        if cell.column == 2:
            cell.number_format = '#,##0'
            cell.alignment = right
        else:
            cell.alignment = center
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 18

# 图表
plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'PingFang HK']
plt.rcParams['axes.unicode_minus'] = False

dates = [r[0] for r in rows]
balances = [r[4] for r in rows]
changes = [r[3] for r in rows]
years = [r[1] for r in rows]

fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 10))
fig.suptitle('存钱记录分析（2023-2026）', fontsize=16, fontweight='bold', y=0.98)

ax1.plot(range(len(dates)), balances, color='#2E75B6', linewidth=2, marker='o', markersize=4)
ax1.fill_between(range(len(dates)), balances, alpha=0.15, color='#2E75B6')
ax1.set_title('累计余额走势', fontsize=13, pad=10)
ax1.set_ylabel('余额（元）', fontsize=11)
ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x/10000:.1f}万'))
ax1.set_xticks(range(len(dates)))
ax1.set_xticklabels([f"{d[5:]}月\n{y}" for d, y in zip(dates, years)], rotation=0, fontsize=7)
ax1.grid(True, alpha=0.3)

colors = ['#70AD47' if c > 0 else '#FF6B6B' for c in changes]
ax2.bar(range(len(dates)), changes, color=colors, alpha=0.8)
ax2.set_title('每笔收支明细', fontsize=13, pad=10)
ax2.set_ylabel('金额（元）', fontsize=11)
ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x/10000:.1f}万'))
ax2.set_xticks(range(len(dates)))
ax2.set_xticklabels([f"{d[5:]}月\n{y}" for d, y in zip(dates, years)], rotation=0, fontsize=7)
ax2.grid(True, alpha=0.3, axis='y')
from matplotlib.patches import Patch
legend_elements = [Patch(facecolor='#70AD47', label='存入'), Patch(facecolor='#FF6B6B', label='支出')]
ax2.legend(handles=legend_elements, loc='upper right')

plt.tight_layout()
chart_path = '/Users/mango/.qclaw/workspace/savings_chart.png'
plt.savefig(chart_path, dpi=150, bbox_inches='tight')
plt.close()

from openpyxl.drawing.image import Image as XLImage
ws3 = wb.create_sheet("图表")
img = XLImage(chart_path)
img.width = 900
img.height = 640
ws3.add_image(img, 'A1')

out_path = '/Users/mango/.qclaw/workspace/存钱记录.xlsx'
wb.save(out_path)
print(f"Done: {out_path}")
print(f"记录数: {len(rows)}, 最终余额: {rows[-1][4]:,}")
print(f"总存入: {total_income:,}, 总支出: {abs(total_expense):,}")
